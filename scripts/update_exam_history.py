"""
update_exam_history.py
KPC 기술사문제검색 엑셀의 '기출' 탭을 data/exam_history.json으로 변환한다.
GitHub Actions(update_exam_data.yml)에서 자동 실행됨.
"""

import sys
import re
import json
import openpyxl

OUTPUT_FILE = "data/exam_history.json"


def parse_round(value) -> int:
    """
    회차 값을 정수로 변환.
    '138', 138, '108회', '제108회' 등 다양한 형식 처리.
    파싱 실패 시 0 반환.
    """
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)

    s = str(value).strip()
    match = re.search(r"\d+", s)
    if match:
        return int(match.group())
    return 0


def main():
    if len(sys.argv) < 2:
        print("사용법: python update_exam_history.py <엑셀파일.xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "기출" not in wb.sheetnames:
        print(f"[ERROR] '기출' 시트를 찾을 수 없음. 시트 목록: {wb.sheetnames}")
        sys.exit(1)

    ws = wb["기출"]

    rows = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        회차, 종목, 유형, 문제 = row[0], row[1], row[2], row[3]
        if 문제 is None:
            continue

        round_no = parse_round(회차)
        if round_no == 0:
            skipped += 1

        rows.append({
            "round":    round_no,
            "subject":  str(종목) if 종목 else "",
            "type":     int(유형) if isinstance(유형, (int, float)) else parse_round(유형),
            "question": str(문제).strip(),
        })

    if skipped:
        print(f"[WARN] 회차 파싱 실패한 행 {skipped}개 (round=0으로 저장됨)")

    rounds = sorted(set(r["round"] for r in rows if r["round"] > 0))
    print(f"총 {len(rows)}개 문제")
    if rounds:
        print(f"회차 범위: {min(rounds)}회 ~ {max(rounds)}회 (총 {len(rounds)}개 회차)")

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=1)

    print(f"저장 완료 → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
